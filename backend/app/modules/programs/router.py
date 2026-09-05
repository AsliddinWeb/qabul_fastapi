from __future__ import annotations

from uuid import UUID

from fastapi import APIRouter, Depends, Query, Request, Response, status
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.dependencies import CurrentUser, get_current_user, get_db, require_permission
from app.core.permissions import Permission
from app.modules.audit.service import AuditService
from app.modules.programs.schemas import (
    BranchCreate,
    BranchRead,
    BranchUpdate,
    EducationFormCreate,
    EducationFormRead,
    EducationFormUpdate,
    EducationLevelCreate,
    EducationLevelRead,
    EducationLevelUpdate,
    ProgramCreate,
    ProgramExpanded,
    ProgramRead,
    ProgramUpdate,
)
from app.modules.programs.service import ProgramsService

router = APIRouter()
require_write = require_permission(Permission.PROGRAMS_WRITE)


def _service(session: AsyncSession = Depends(get_db)) -> ProgramsService:
    return ProgramsService(session)


# ---------- Branches ----------
@router.get("/branches", response_model=list[BranchRead])
async def list_branches(
    active_only: bool = True,
    svc: ProgramsService = Depends(_service),
) -> list[BranchRead]:
    rows = await svc.list_branches(active_only=active_only)
    return [BranchRead.model_validate(r) for r in rows]


@router.post(
    "/branches",
    response_model=BranchRead,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(require_write)],
)
async def create_branch(payload: BranchCreate, svc: ProgramsService = Depends(_service)) -> BranchRead:
    obj = await svc.create_branch(payload)
    await svc.session.commit()
    return BranchRead.model_validate(obj)


@router.patch(
    "/branches/{branch_id}",
    response_model=BranchRead,
    dependencies=[Depends(require_write)],
)
async def update_branch(
    branch_id: UUID,
    payload: BranchUpdate,
    svc: ProgramsService = Depends(_service),
) -> BranchRead:
    obj = await svc.update_branch(branch_id, payload)
    await svc.session.commit()
    return BranchRead.model_validate(obj)


@router.delete(
    "/branches/{branch_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(require_write)],
)
async def delete_branch(branch_id: UUID, svc: ProgramsService = Depends(_service)) -> None:
    await svc.delete_branch(branch_id)
    await svc.session.commit()


# ---------- Education Levels ----------
@router.get("/education-levels", response_model=list[EducationLevelRead])
async def list_education_levels(svc: ProgramsService = Depends(_service)) -> list[EducationLevelRead]:
    rows = await svc.list_education_levels()
    return [EducationLevelRead.model_validate(r) for r in rows]


@router.post(
    "/education-levels",
    response_model=EducationLevelRead,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(require_write)],
)
async def create_education_level(
    payload: EducationLevelCreate,
    svc: ProgramsService = Depends(_service),
) -> EducationLevelRead:
    obj = await svc.create_education_level(payload)
    await svc.session.commit()
    return EducationLevelRead.model_validate(obj)


@router.patch(
    "/education-levels/{level_id}",
    response_model=EducationLevelRead,
    dependencies=[Depends(require_write)],
)
async def update_education_level(
    level_id: UUID,
    payload: EducationLevelUpdate,
    svc: ProgramsService = Depends(_service),
) -> EducationLevelRead:
    obj = await svc.update_education_level(level_id, payload)
    await svc.session.commit()
    return EducationLevelRead.model_validate(obj)


@router.delete(
    "/education-levels/{level_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(require_write)],
)
async def delete_education_level(level_id: UUID, svc: ProgramsService = Depends(_service)) -> None:
    await svc.delete_education_level(level_id)
    await svc.session.commit()


# ---------- Education Forms ----------
@router.get("/education-forms", response_model=list[EducationFormRead])
async def list_education_forms(svc: ProgramsService = Depends(_service)) -> list[EducationFormRead]:
    rows = await svc.list_education_forms()
    return [EducationFormRead.model_validate(r) for r in rows]


@router.post(
    "/education-forms",
    response_model=EducationFormRead,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(require_write)],
)
async def create_education_form(
    payload: EducationFormCreate,
    svc: ProgramsService = Depends(_service),
) -> EducationFormRead:
    obj = await svc.create_education_form(payload)
    await svc.session.commit()
    return EducationFormRead.model_validate(obj)


@router.patch(
    "/education-forms/{form_id}",
    response_model=EducationFormRead,
    dependencies=[Depends(require_write)],
)
async def update_education_form(
    form_id: UUID,
    payload: EducationFormUpdate,
    svc: ProgramsService = Depends(_service),
) -> EducationFormRead:
    obj = await svc.update_education_form(form_id, payload)
    await svc.session.commit()
    return EducationFormRead.model_validate(obj)


@router.delete(
    "/education-forms/{form_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(require_write)],
)
async def delete_education_form(form_id: UUID, svc: ProgramsService = Depends(_service)) -> None:
    await svc.delete_education_form(form_id)
    await svc.session.commit()


# ---------- Programs ----------
@router.get("/programs", response_model=list[ProgramExpanded])
async def list_programs(
    active_only: bool = True,
    branch_id: UUID | None = Query(default=None),
    svc: ProgramsService = Depends(_service),
) -> list[ProgramExpanded]:
    rows = await svc.list_programs_expanded(active_only=active_only, branch_id=branch_id)
    return [ProgramExpanded.model_validate(r) for r in rows]


@router.get(
    "/programs/export.xlsx",
    dependencies=[Depends(require_permission(Permission.PROGRAMS_READ))],
)
async def export_programs_xlsx(
    active_only: bool = False,
    branch_id: UUID | None = Query(default=None),
    education_level_id: UUID | None = Query(default=None),
    education_form_id: UUID | None = Query(default=None),
    search: str | None = Query(default=None, max_length=100),
    request: Request = None,  # type: ignore[assignment]
    current: CurrentUser = Depends(get_current_user),
    svc: ProgramsService = Depends(_service),
) -> Response:
    """Export the programs catalog to Excel, honoring the list-page filters."""
    import io
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows = await svc.list_programs_expanded(active_only=active_only, branch_id=branch_id)
    if education_level_id is not None:
        rows = [r for r in rows if str(r.get("education_level_id")) == str(education_level_id)]
    if education_form_id is not None:
        rows = [r for r in rows if str(r.get("education_form_id")) == str(education_form_id)]
    if search:
        q = search.strip().lower()
        rows = [
            r for r in rows
            if q in (r.get("name") or "").lower()
            or q in (r.get("code") or "").lower()
            or q in (r.get("branch_name") or "").lower()
        ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Yo'nalishlar"
    headers = [
        "№", "Yo'nalish", "Kodi", "Filial", "Daraja", "Ta'lim shakli",
        "Kontrakt narxi", "O'qish muddati (yil)", "Shartnoma seriyasi", "Holati",
    ]
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    ws.append(headers)
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    for i, r in enumerate(rows, start=1):
        ws.append([
            i,
            r.get("name") or "",
            r.get("code") or "",
            r.get("branch_name") or "",
            r.get("education_level_name") or "",
            r.get("education_form_name") or "",
            float(r["tuition_fee"]) if r.get("tuition_fee") is not None else 0,
            r.get("study_duration_years") or "",
            r.get("contract_series") or "",
            "Faol" if r.get("is_active") else "Faol emas",
        ])

    widths = [5, 42, 14, 22, 14, 16, 16, 12, 18, 10]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    for row_cells in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row_cells:
            cell.number_format = "#,##0"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"yonalishlar-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"

    try:
        await AuditService(svc.session).log(
            "programs.export",
            user_id=UUID(current.user_id),
            entity_type="programs",
            changes={"count": len(rows)},
            request=request,
        )
        await svc.session.commit()
    except Exception:  # noqa: BLE001 — never fail the download over an audit hiccup
        pass

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


@router.get("/programs/{program_id}", response_model=ProgramRead)
async def get_program(program_id: UUID, svc: ProgramsService = Depends(_service)) -> ProgramRead:
    obj = await svc.get_program(program_id)
    return ProgramRead.model_validate(obj)


@router.post(
    "/programs",
    response_model=ProgramRead,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(require_write)],
)
async def create_program(payload: ProgramCreate, svc: ProgramsService = Depends(_service)) -> ProgramRead:
    obj = await svc.create_program(payload)
    await svc.session.commit()
    return ProgramRead.model_validate(obj)


@router.patch(
    "/programs/{program_id}",
    response_model=ProgramRead,
    dependencies=[Depends(require_write)],
)
async def update_program(
    program_id: UUID,
    payload: ProgramUpdate,
    svc: ProgramsService = Depends(_service),
) -> ProgramRead:
    obj = await svc.update_program(program_id, payload)
    await svc.session.commit()
    return ProgramRead.model_validate(obj)


@router.delete(
    "/programs/{program_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(require_write)],
)
async def delete_program(program_id: UUID, svc: ProgramsService = Depends(_service)) -> None:
    await svc.delete_program(program_id)
    await svc.session.commit()
