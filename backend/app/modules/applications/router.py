from __future__ import annotations

import csv
import io
from datetime import datetime
from enum import Enum
from uuid import UUID


from fastapi import APIRouter, BackgroundTasks, Depends, Query, Request, Response, status
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.dependencies import (
    CurrentUser,
    get_current_user,
    get_db,
    operator_scope,
    require_export_access,
    require_permission,
    require_root_superadmin,
)
from app.core.exceptions import ForbiddenError, NotFoundError
from app.core.permissions import Permission
from app.core.schemas import PageResponse
from app.db.enums import AdmissionType, ApplicationStatus, ContractStatus, Gender
from app.integrations.crm.events import enqueue_application_status_event
from app.integrations.telegram.notifier import enqueue_application_created
from app.modules.applicants.repository import ApplicantRepository
from app.modules.applications.schemas import (
    ApplicationCreateForApplicant,
    ApplicationCreateSelf,
    ApplicationDetailed,
    ApplicationRead,
    ApplicationReview,
    ApplicationUpdate,
    HemisDecision,
    ReassignOperator,
)
from app.modules.applications.service import ApplicationsService
from app.modules.audit.service import AuditService

router = APIRouter()


def _service(session: AsyncSession = Depends(get_db)) -> ApplicationsService:
    return ApplicationsService(session)


# ---------- Self ----------
@router.get(
    "/me",
    response_model=list[ApplicationDetailed],
    dependencies=[Depends(require_permission(Permission.APPLICATIONS_READ_SELF))],
)
async def list_my_applications(
    current: CurrentUser = Depends(get_current_user),
    svc: ApplicationsService = Depends(_service),
) -> list[ApplicationDetailed]:
    applicants = ApplicantRepository(svc.session)
    me = await applicants.get_by_user_id(UUID(current.user_id))
    if not me:
        return []
    rows = await svc.list_detailed_for_applicant(me.id)
    return [ApplicationDetailed.model_validate(r) for r in rows]


@router.post(
    "/me",
    response_model=ApplicationRead,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(require_permission(Permission.APPLICATIONS_CREATE_SELF))],
)
async def submit_my_application(
    payload: ApplicationCreateSelf,
    request: Request,
    background: BackgroundTasks,
    current: CurrentUser = Depends(get_current_user),
    svc: ApplicationsService = Depends(_service),
) -> ApplicationRead:
    applicants = ApplicantRepository(svc.session)
    me = await applicants.get_by_user_id(UUID(current.user_id))
    if not me:
        raise ForbiddenError("Complete your applicant profile before submitting an application")

    obj = await svc.create(
        applicant_id=me.id,
        payload=payload,
        actor_id=UUID(current.user_id),
    )
    await AuditService(svc.session).log(
        "application.create",
        user_id=UUID(current.user_id),
        entity_type="applications",
        entity_id=obj.id,
        changes={
            "applicant_id": str(me.id),
            "program_id": str(payload.program_id),
            "admission_type": payload.admission_type.value,
        },
        request=request,
    )
    await svc.session.commit()

    enqueue_application_status_event(
        background,
        external_id=str(me.id),
        status=obj.status.value,
        note=f"submitted application {obj.application_number}",
    )
    return ApplicationRead.model_validate(obj)


@router.post(
    "/me/{application_id}/withdraw",
    response_model=ApplicationRead,
    dependencies=[Depends(require_permission(Permission.APPLICATIONS_WITHDRAW_SELF))],
)
async def withdraw_my_application(
    application_id: UUID,
    request: Request,
    background: BackgroundTasks,
    current: CurrentUser = Depends(get_current_user),
    svc: ApplicationsService = Depends(_service),
) -> ApplicationRead:
    applicants = ApplicantRepository(svc.session)
    me = await applicants.get_by_user_id(UUID(current.user_id))
    if not me:
        raise ForbiddenError("No applicant profile")

    await svc.get_or_403_for_applicant(application_id, applicant_id=me.id)
    obj = await svc.withdraw(application_id, actor_id=UUID(current.user_id))
    await AuditService(svc.session).log(
        "application.withdraw",
        user_id=UUID(current.user_id),
        entity_type="applications",
        entity_id=obj.id,
        request=request,
    )
    await svc.session.commit()

    enqueue_application_status_event(
        background,
        external_id=str(me.id),
        status=obj.status.value,
        note="withdrawn by applicant",
    )
    return ApplicationRead.model_validate(obj)


# ---------- Staff ----------
@router.get(
    "/stats",
    response_model=dict[str, int],
    dependencies=[Depends(require_permission(Permission.APPLICATIONS_LIST))],
)
async def application_stats(svc: ApplicationsService = Depends(_service)) -> dict[str, int]:
    return await svc.status_counts()


@router.get(
    "/trend",
    response_model=list[dict],
    dependencies=[Depends(require_permission(Permission.APPLICATIONS_LIST))],
)
async def application_trend(
    months: int = Query(default=12, ge=1, le=24),
    svc: ApplicationsService = Depends(_service),
) -> list[dict]:
    """Monthly application counts (per-status), 12 buckets by default. For dashboard chart."""
    return await svc.monthly_trend(months)


@router.get(
    "",
    response_model=PageResponse[ApplicationDetailed],
    dependencies=[Depends(require_permission(Permission.APPLICATIONS_LIST))],
)
async def list_applications(
    status_filter: ApplicationStatus | None = Query(default=None, alias="status"),
    admission_type: AdmissionType | None = Query(default=None),
    program_id: UUID | None = Query(default=None),
    branch_id: UUID | None = Query(default=None),
    education_level_id: UUID | None = Query(default=None),
    education_form_id: UUID | None = Query(default=None),
    consulting_agency_id: UUID | None = Query(default=None),
    registered_by_id: UUID | None = Query(default=None, description="Filter by who registered the applicant (operator attribution)"),
    source: str | None = Query(default=None, description="'lead' / 'direct' / 'self' — origin of the application"),
    search: str | None = Query(default=None, max_length=100, description="Fuzzy match on application_number, F.I.Sh., passport, PINFL"),
    created_from: datetime | None = Query(default=None, description="created_at >= this UTC timestamp"),
    created_to: datetime | None = Query(default=None, description="created_at <= this UTC timestamp"),
    page: int = Query(default=1, ge=1),
    size: int = Query(default=20, ge=1, le=100),
    current: CurrentUser = Depends(get_current_user),
    svc: ApplicationsService = Depends(_service),
) -> PageResponse[ApplicationDetailed]:
    # Operators are confined to applicants they registered; other roles see all.
    scope = operator_scope(current)
    if scope is not None:
        registered_by_id = scope
    items, total = await svc.list_detailed(
        status=status_filter,
        admission_type=admission_type,
        program_id=program_id,
        branch_id=branch_id,
        education_level_id=education_level_id,
        education_form_id=education_form_id,
        consulting_agency_id=consulting_agency_id,
        registered_by_id=registered_by_id,
        source=source,
        search=search,
        created_from=created_from,
        created_to=created_to,
        limit=size,
        offset=(page - 1) * size,
    )
    return PageResponse[ApplicationDetailed].build(
        items=[ApplicationDetailed.model_validate(a) for a in items],
        total=total, page=page, size=size,
    )


@router.get(
    "/export.csv",
    # Security: export is limited to the ROOT superadmin OR a role explicitly
    # granted applications.export (e.g. accountants — revocable per user).
    # Every other role (incl. other superadmins/admins) gets 403. Audit-logged.
    dependencies=[Depends(require_export_access(Permission.APPLICATIONS_EXPORT))],
)
async def export_applications_csv(
    status_filter: ApplicationStatus | None = Query(default=None, alias="status"),
    admission_type: AdmissionType | None = Query(default=None),
    program_id: UUID | None = Query(default=None),
    branch_id: UUID | None = Query(default=None),
    education_level_id: UUID | None = Query(default=None),
    education_form_id: UUID | None = Query(default=None),
    consulting_agency_id: UUID | None = Query(default=None),
    registered_by_id: UUID | None = Query(default=None),
    source: str | None = Query(default=None),
    search: str | None = Query(default=None, max_length=100),
    created_from: datetime | None = Query(default=None),
    created_to: datetime | None = Query(default=None),
    request: Request = None,  # type: ignore[assignment]
    current: CurrentUser = Depends(get_current_user),
    svc: ApplicationsService = Depends(_service),
) -> Response:
    """Export filtered applications to CSV."""
    items, _ = await svc.list_detailed(
        status=status_filter,
        admission_type=admission_type,
        program_id=program_id,
        branch_id=branch_id,
        education_level_id=education_level_id,
        education_form_id=education_form_id,
        consulting_agency_id=consulting_agency_id,
        registered_by_id=registered_by_id,
        source=source,
        search=search,
        created_from=created_from,
        created_to=created_to,
        limit=10_000,
        offset=0,
    )
    buf = io.StringIO()
    buf.write("﻿")
    w = csv.writer(buf)
    w.writerow([
        "Ariza raqami", "Holati", "Topshirish turi", "F.I.Sh.", "Yo'nalish",
        "Filial", "Bosqich", "Shakl", "Konsalting", "Yaratilgan", "Topshirilgan",
    ])
    for a in items:
        st = a.get("status")
        at = a.get("admission_type")
        w.writerow([
            a.get("application_number") or "",
            (st.value if hasattr(st, "value") else st or ""),
            (at.value if hasattr(at, "value") else at or ""),
            a.get("applicant_full_name") or "",
            a.get("program_name") or "",
            a.get("branch_name") or "",
            a.get("education_level_name") or "",
            a.get("education_form_name") or "",
            a.get("consulting_agency_name") or "",
            a["created_at"].strftime("%Y-%m-%d %H:%M") if a.get("created_at") else "",
            a["submitted_at"].strftime("%Y-%m-%d %H:%M") if a.get("submitted_at") else "",
        ])
    fn = f"applications-{datetime.now().strftime('%Y%m%d-%H%M')}.csv"

    # Record WHO exported and HOW MANY rows — surfaced in the audit log
    # (action "applications.export") so downloads are attributable per user.
    try:
        await AuditService(svc.session).log(
            "applications.export",
            user_id=UUID(current.user_id),
            entity_type="applications",
            changes={
                "rows": len(items),
                "status": status_filter.value if status_filter else None,
                "program_id": str(program_id) if program_id else None,
                "branch_id": str(branch_id) if branch_id else None,
                "registered_by_id": str(registered_by_id) if registered_by_id else None,
            },
            request=request,
        )
        await svc.session.commit()
    except Exception:
        pass  # never fail the download over an audit write

    return Response(
        content=buf.getvalue().encode("utf-8"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


# Enum → Uzbek label map. Used by _xlsx_cell_value so the workbook reads
# "Qabul qilindi" instead of "ApplicationStatus.ACCEPTED" or "qabul_qilindi".
# Add a new enum here whenever a new enum column lands in _XLSX_COLUMNS.
_ENUM_LABELS: dict = {
    ApplicationStatus.PENDING:  "Topshirildi",
    ApplicationStatus.REVIEW:   "Ko'rib chiqilmoqda",
    ApplicationStatus.ACCEPTED: "Qabul qilindi",
    ApplicationStatus.REJECTED: "Rad etildi",

    AdmissionType.REGULAR:      "Yangi qabul (1-kurs)",
    AdmissionType.TRANSFER:     "O'qishni ko'chirish",
    AdmissionType.SECOND_SPEC:  "2-mutaxassislik",
    AdmissionType.MAGISTRATURA: "Magistratura",

    Gender.MALE:   "Erkak",
    Gender.FEMALE: "Ayol",

    ContractStatus.DRAFT:     "Loyiha",
    ContractStatus.SIGNED:    "Imzolangan",
    ContractStatus.CANCELLED: "Bekor qilingan",
    ContractStatus.COMPLETED: "Yakunlangan",
}

# Source string (set in repository.list_for_export) → Uzbek label.
_SOURCE_LABELS: dict[str, str] = {
    "lead":   "Lead'dan",
    "direct": "Operator kiritgan",
    "self":   "Saytdan o'zi",
}

# Human-readable labels used as XLSX header row. Order drives the
# column order in the file. Layout follows the university's existing
# Excel template (the one staff already use):
#
#   Block 1 — Shartnoma raqami → Shartnoma sanasi → F.I.Sh.
#             (the 3 cells that drive every search-by-contract or
#             search-by-name flow; placed first so they're on screen
#             without horizontal scrolling)
#   Block 2 — Pasport, PINFL, Guruhi (group code), Telefon, Filial,
#             Ta'lim yo'nalishi, Ta'lim shakli, Shartnoma summasi, Kurs
#             (matches the legacy template column order verbatim)
#   Block 3 — O'qishni ko'chirish info — country + prior university +
#             prior diplom + applicant's diplom (1-kurs OR Bakalavr)
#   Block 4 — Remaining applicant identity / contact / geo
#   Block 5 — Misc application meta (status, manba, operator, …)
#   Block 6 — Timeline + free-text + UUIDs at the very end
_XLSX_COLUMNS: list[tuple[str, str]] = [
    # -- Block 1: Top-of-mind --
    ("contract_number",         "Shartnoma №"),
    ("contract_signed_at",      "Shartnoma sanasi"),
    ("applicant_full_name",     "F.I.Sh."),

    # -- Block 2: Legacy template columns, in the SAME order the staff
    #             already use day-to-day --
    ("passport_series",         "Pasport"),
    ("pinfl",                   "PINFL"),
    ("guruhi",                  "Guruhi"),
    ("phone",                   "Telefon"),
    ("branch_name",             "Fakultet / Filial"),
    ("program_name",            "Ta'lim yo'nalishi"),
    ("education_form_name",     "Ta'lim shakli"),
    ("contract_total_amount",   "Shartnoma summasi"),
    ("course_label",            "Kurs"),

    # -- Block 3: O'qishni ko'chirish (perevod) source + diplom info --
    ("transfer_country_name",   "Ko'chirish: Davlat"),
    ("transfer_university_name", "Ko'chirish: OTM nomi"),
    ("diplom_serial_number",    "Diplom №"),
    ("diplom_university_name",  "Diplom OTM"),
    ("diplom_graduation_year",  "Bitirgan yil"),
    ("diplom_region_name",      "Diplom viloyati"),
    ("diplom_district_name",    "Diplom tumani"),

    # -- Block 4: Remaining identity / contact / geo --
    ("last_name",               "Familiya"),
    ("first_name",              "Ism"),
    ("other_name",              "Otasining ismi"),
    ("birth_date",              "Tug'ilgan sana"),
    ("gender",                  "Jinsi"),
    ("nationality",             "Millati"),
    ("telegram_username",       "Telegram"),
    ("additional_phone",        "Qo'shimcha telefon"),
    ("region_name",             "Viloyat"),
    ("district_name",           "Tuman"),
    ("address",                 "Manzil"),

    # -- Block 5: Application meta --
    ("application_number",      "Ariza №"),
    ("status",                  "Ariza holati"),
    ("admission_type",          "Qabul turi"),
    ("source",                  "Manba"),
    ("lead_source_code",        "Lead manba kodi"),
    ("program_code",            "Yo'nalish kodi"),
    ("education_level_name",    "Ta'lim darajasi"),
    ("program_tuition_fee",     "Yo'nalish summasi"),
    ("operator_full_name",      "Operator"),
    ("consulting_agency_name",  "Konsalting agentligi"),
    ("contract_status",         "Shartnoma holati"),

    # -- Block 6: Timeline + free-text + UUIDs --
    ("created_at",              "Yaratilgan"),
    ("submitted_at",            "Topshirilgan"),
    ("reviewed_at",             "Ko'rib chiqilgan"),
    ("rejection_reason",        "Rad etish sababi"),
    ("notes",                   "Izoh"),
    ("application_id",          "Ariza UUID"),
    ("applicant_id",            "Abituriyent UUID"),
]


def _xlsx_cell_value(key: str, value):
    """Coerce repo values into XLSX-friendly types.

    Enums → Uzbek label via _ENUM_LABELS (falls back to .value if missing,
    so a newly-added enum still exports — just untranslated). Source
    string ("lead"/"direct") → Uzbek label. UUID → str (openpyxl raises
    ValueError on UUID). Datetime → tz-stripped (openpyxl can't store
    tz). None → empty cell. Decimal and date pass through natively.
    """
    if value is None:
        return ""
    if isinstance(value, Enum):
        return _ENUM_LABELS.get(value, value.value)
    if key == "source" and isinstance(value, str):
        return _SOURCE_LABELS.get(value, value)
    if isinstance(value, UUID):
        return str(value)
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


@router.get(
    "/export.xlsx",
    # Root superadmin OR a role granted applications.export (accountants —
    # revocable per user). Every other role gets 403. Audit-logged.
    dependencies=[Depends(require_export_access(Permission.APPLICATIONS_EXPORT))],
)
async def export_applications_xlsx(
    status_filter: ApplicationStatus | None = Query(default=None, alias="status"),
    admission_type: AdmissionType | None = Query(default=None),
    program_id: UUID | None = Query(default=None),
    branch_id: UUID | None = Query(default=None),
    education_level_id: UUID | None = Query(default=None),
    education_form_id: UUID | None = Query(default=None),
    consulting_agency_id: UUID | None = Query(default=None),
    registered_by_id: UUID | None = Query(default=None),
    source: str | None = Query(default=None),
    search: str | None = Query(default=None, max_length=100),
    created_from: datetime | None = Query(default=None),
    created_to: datetime | None = Query(default=None),
    request: Request = None,  # type: ignore[assignment]
    current: CurrentUser = Depends(get_current_user),
    svc: ApplicationsService = Depends(_service),
) -> Response:
    """Export filtered applications to a styled Excel (.xlsx) workbook.

    Heavy lift in three places:
      1. Repository.list_for_export does ONE query with all the joins
         (region/district/contract/operator) so the endpoint is a thin
         formatting layer, not a per-row N+1.
      2. Header row gets a brand-coloured fill + bold white text + frozen
         pane so it stays visible as the user scrolls 388 rows.
      3. Column widths are sized from the actual content rather than a
         fixed value — long phone numbers and F.I.Sh. strings would
         otherwise spill or truncate visibly in Excel's default 8.43ch.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows = await svc.list_for_export(
        status=status_filter,
        admission_type=admission_type,
        program_id=program_id,
        branch_id=branch_id,
        education_level_id=education_level_id,
        education_form_id=education_form_id,
        consulting_agency_id=consulting_agency_id,
        registered_by_id=registered_by_id,
        source=source,
        search=search,
        created_from=created_from,
        created_to=created_to,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Arizalar"

    header_fill = PatternFill("solid", fgColor="4F46E5")  # indigo-600 — brand
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Header row
    headers = [label for _, label in _XLSX_COLUMNS]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"  # keep header visible on scroll
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Data rows
    keys = [k for k, _ in _XLSX_COLUMNS]
    for row in rows:
        ws.append([_xlsx_cell_value(k, row.get(k)) for k in keys])

    # Column widths — fit to the longest value in each column, capped at 50
    # so a stray 500-char "notes" doesn't blow up the whole sheet.
    for col_idx, key in enumerate(keys, start=1):
        max_len = len(headers[col_idx - 1])
        for row in rows:
            v = row.get(key)
            if v is None:
                continue
            s = str(v.value) if hasattr(v, "value") and not isinstance(v, (str, int, float, bool)) else str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    # Row banding via alternating fill — easier to follow a single row across
    # 35+ columns. openpyxl has no built-in zebra so we set it per cell;
    # cheap enough at this row count.
    band_fill = PatternFill("solid", fgColor="F8FAFC")  # slate-50
    for r in range(2, ws.max_row + 1):
        if r % 2 == 0:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = band_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Filename: arizalar_2026-06-11_17-35.xlsx — readable date + time so
    # users can tell exports apart without opening each file. Underscores
    # because Windows hides hyphens and colons confuse the OS.
    fn = f"arizalar_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

    # Attribute the export in the audit log (action "applications.export").
    try:
        await AuditService(svc.session).log(
            "applications.export",
            user_id=UUID(current.user_id),
            entity_type="applications",
            changes={
                "format": "xlsx",
                "rows": len(rows),
                "status": status_filter.value if status_filter else None,
                "program_id": str(program_id) if program_id else None,
                "branch_id": str(branch_id) if branch_id else None,
                "registered_by_id": str(registered_by_id) if registered_by_id else None,
            },
            request=request,
        )
        await svc.session.commit()
    except Exception:
        pass  # never fail the download over an audit write

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


@router.post(
    "",
    response_model=ApplicationRead,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(require_permission(Permission.APPLICATIONS_REVIEW))],
)
async def staff_create_application(
    payload: ApplicationCreateForApplicant,
    request: Request,
    bg: BackgroundTasks,
    current: CurrentUser = Depends(get_current_user),
    svc: ApplicationsService = Depends(_service),
) -> ApplicationRead:
    obj = await svc.create_for_applicant(payload, actor_id=UUID(current.user_id))

    # If this application is the conversion of a Lead, mark it as won + link back.
    if payload.lead_id:
        from app.modules.leads.service import LeadService
        from app.modules.leads.repository import LeadSourceRepository
        # Stamp source code on the application for analytics (denormalised cache).
        lead_svc = LeadService(svc.session)
        lead = await lead_svc.leads.get(payload.lead_id)
        if lead:
            obj.lead_id = lead.id
            if lead.source_id:
                src = await LeadSourceRepository(svc.session).get(lead.source_id)
                obj.lead_source_code = src.code if src else None
            await lead_svc.mark_converted(
                lead.id,
                applicant_id=payload.applicant_id,
                application_id=obj.id,
                actor_id=UUID(current.user_id),
            )

    await AuditService(svc.session).log(
        "application.create_by_staff",
        user_id=UUID(current.user_id),
        entity_type="applications",
        entity_id=obj.id,
        changes={
            "applicant_id": str(payload.applicant_id),
            "program_id": str(payload.program_id),
            "admission_type": payload.admission_type.value,
            "lead_id": str(payload.lead_id) if payload.lead_id else None,
        },
        request=request,
    )
    await svc.session.commit()

    # Push the new application to the Telegram notification bot (best-effort,
    # runs after the response is sent).
    payload_dict = await svc.notification_payload(obj.id)
    if payload_dict:
        enqueue_application_created(bg, payload=payload_dict)

    return ApplicationRead.model_validate(obj)


@router.post(
    "/{application_id}/hemis",
    response_model=ApplicationRead,
    dependencies=[Depends(require_permission(Permission.APPLICATIONS_REVIEW))],
)
async def set_hemis_status(
    application_id: UUID,
    payload: HemisDecision,
    request: Request,
    current: CurrentUser = Depends(get_current_user),
    svc: ApplicationsService = Depends(_service),
) -> ApplicationRead:
    """Toggle an application's HEMIS enrolment state — called by the Telegram
    bot when an operator presses ✅ (qoshildi) / ❌ (qoshilmadi)."""
    obj = await svc.set_hemis_status(
        application_id, status=payload.status, marked_by=payload.marked_by, comment=payload.comment,
    )
    await AuditService(svc.session).log(
        f"application.hemis_{payload.status}",
        user_id=UUID(current.user_id),
        entity_type="applications",
        entity_id=obj.id,
        changes={"hemis_status": payload.status, "marked_by": payload.marked_by, "comment": payload.comment},
        request=request,
    )
    await svc.session.commit()
    return ApplicationRead.model_validate(obj)


@router.post(
    "/{application_id}/reassign-operator",
    response_model=ApplicationRead,
    dependencies=[Depends(require_root_superadmin)],
)
async def reassign_operator(
    application_id: UUID,
    payload: ReassignOperator,
    request: Request,
    current: CurrentUser = Depends(get_current_user),
    svc: ApplicationsService = Depends(_service),
) -> ApplicationRead:
    """Root-superadmin only: change the operator credited with this application
    (reassigns the applicant's registered_by_id)."""
    obj = await svc.reassign_operator(application_id, operator_id=payload.operator_id)
    await AuditService(svc.session).log(
        "application.reassign_operator",
        user_id=UUID(current.user_id),
        entity_type="applications",
        entity_id=obj.id,
        changes={"operator_id": str(payload.operator_id), "applicant_id": str(obj.applicant_id)},
        request=request,
    )
    await svc.session.commit()
    return ApplicationRead.model_validate(obj)


@router.get(
    "/{application_id}",
    response_model=ApplicationRead,
    dependencies=[Depends(require_permission(Permission.APPLICATIONS_READ))],
)
async def get_application(
    application_id: UUID,
    current: CurrentUser = Depends(get_current_user),
    svc: ApplicationsService = Depends(_service),
) -> ApplicationRead:
    obj = await svc.get(application_id)
    # Operators may only open applications of applicants they registered.
    scope = operator_scope(current)
    if scope is not None:
        applicant = await svc.applicants.get(obj.applicant_id)
        if applicant is None or applicant.registered_by_id != scope:
            raise NotFoundError("Application not found")
    return ApplicationRead.model_validate(obj)


@router.patch(
    "/{application_id}",
    response_model=ApplicationRead,
    dependencies=[Depends(require_permission(Permission.APPLICATIONS_REVIEW))],
)
async def staff_update_application(
    application_id: UUID,
    payload: ApplicationUpdate,
    request: Request,
    current: CurrentUser = Depends(get_current_user),
    svc: ApplicationsService = Depends(_service),
) -> ApplicationRead:
    obj = await svc.update(application_id, payload)
    await AuditService(svc.session).log(
        "application.update",
        user_id=UUID(current.user_id),
        entity_type="applications",
        entity_id=obj.id,
        changes=payload.model_dump(exclude_unset=True, mode="json"),
        request=request,
    )
    await svc.session.commit()
    return ApplicationRead.model_validate(obj)


@router.delete(
    "/{application_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(require_permission(Permission.APPLICATIONS_REVIEW))],
)
async def staff_delete_application(
    application_id: UUID,
    request: Request,
    current: CurrentUser = Depends(get_current_user),
    svc: ApplicationsService = Depends(_service),
):
    # Snapshot key fields BEFORE delete so the audit row carries a useful
    # body — once the application is gone we can't read them back. We also
    # capture the applicant's full name + phone so the audit detail page
    # can show them without needing to hit the (possibly also-deleted)
    # applicant row later.
    from sqlalchemy import select as _select
    from app.modules.applicants.models import Applicant as _Applicant
    from app.modules.users.models import User as _User

    app = await svc.get(application_id)
    snapshot: dict[str, object | None] = {
        "application_number": app.application_number,
        "applicant_id": str(app.applicant_id),
        "status": app.status.value if app.status else None,
    }
    ap_row = (await svc.session.execute(
        _select(_Applicant.last_name, _Applicant.first_name, _Applicant.other_name,
                _Applicant.additional_phone, _User.phone)
        .join(_User, _User.id == _Applicant.user_id)
        .where(_Applicant.id == app.applicant_id)
    )).first()
    if ap_row is not None:
        name_parts = [ap_row.last_name, ap_row.first_name, ap_row.other_name]
        snapshot["applicant_full_name"] = " ".join(p for p in name_parts if p)
        snapshot["applicant_phone"] = ap_row.phone or ap_row.additional_phone
    await svc.delete(application_id)
    await AuditService(svc.session).log(
        "application.delete",
        user_id=UUID(current.user_id),
        entity_type="applications",
        entity_id=application_id,
        changes=snapshot,
        request=request,
    )
    await svc.session.commit()


@router.post(
    "/bulk-delete",
    dependencies=[Depends(require_permission(Permission.APPLICATIONS_REVIEW))],
)
async def bulk_delete_applications(
    payload: dict,
    svc: ApplicationsService = Depends(_service),
) -> dict:
    """Bulk-delete applications. Lighter than the per-row endpoint —
    skips the rich audit snapshot for each row because writing N
    audit_logs on a bulk action floods the timeline. The action itself
    is unmistakable; if forensic detail per row is needed, do them one
    by one.
    """
    ids_raw = payload.get("ids") or []
    if not isinstance(ids_raw, list) or not ids_raw:
        raise HTTPException(status_code=400, detail="ids list is required")
    try:
        ids = [UUID(str(i)) for i in ids_raw]
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Invalid id in list")
    deleted = 0
    skipped = 0
    for app_id in ids:
        try:
            await svc.delete(app_id)
            deleted += 1
        except Exception:
            skipped += 1
    await svc.session.commit()
    return {"deleted": deleted, "skipped": skipped}


@router.post(
    "/{application_id}/review",
    response_model=ApplicationRead,
    dependencies=[Depends(require_permission(Permission.APPLICATIONS_REVIEW))],
)
async def review_application(
    application_id: UUID,
    payload: ApplicationReview,
    request: Request,
    background: BackgroundTasks,
    current: CurrentUser = Depends(get_current_user),
    svc: ApplicationsService = Depends(_service),
) -> ApplicationRead:
    obj = await svc.review(application_id, payload, reviewer_id=UUID(current.user_id))
    await AuditService(svc.session).log(
        "application.review",
        user_id=UUID(current.user_id),
        entity_type="applications",
        entity_id=obj.id,
        changes={
            "approved": payload.approved,
            "rejection_reason": payload.rejection_reason,
        },
        request=request,
    )
    await svc.session.commit()

    enqueue_application_status_event(
        background,
        external_id=str(obj.applicant_id),
        status=obj.status.value,
        note=payload.rejection_reason if not payload.approved else "approved",
    )
    return ApplicationRead.model_validate(obj)


@router.post(
    "/{application_id}/start-review",
    response_model=ApplicationRead,
    dependencies=[Depends(require_permission(Permission.APPLICATIONS_REVIEW))],
)
async def start_review(
    application_id: UUID,
    current: CurrentUser = Depends(get_current_user),
    svc: ApplicationsService = Depends(_service),
) -> ApplicationRead:
    obj = await svc.mark_review(application_id, reviewer_id=UUID(current.user_id))
    await svc.session.commit()
    return ApplicationRead.model_validate(obj)
